# coding=utf-8

import os
import sys

import openpyxl

# 添加工程路径。防止找不到模块
base_path = os.getcwd()
sys.path.append(base_path)
# 获取当前文件路径，用工程路径调用是会出错。
t_path = os.path.abspath(os.path.dirname(__file__))


class Read_Excel:
    def __init__(self):
        self.open_excel = openpyxl.load_workbook(t_path + '\\excel_data\\excel_data.xlsx')
        # 拿到所有
        self.sheet = self.open_excel.sheetnames

    def get_sheet_data(self, index=0):
        '''

        :param index: 第几个sheet，从0开始算
        :return: 返回整个表的值
        '''
        excel_data = self.open_excel[self.sheet[index]]
        return excel_data

    def get_data(self, row, col):
        '''

        :param row: 行数
        :param col: 列数
        :return: 返回的是指定行数和列数的单元格的值
        '''
        date = self.get_sheet_data().cell(row=row, column=col).value
        return date

    def get_row(self):
        '''

        :return: 返回一个最大行数值
        '''
        row_max = self.get_sheet_data().max_row
        return row_max

    def get_data_row(self, row):
        '''

        :param row: 行数
        :return: 返回一整行的数据
        '''
        # 获取某一行的数据
        # 从一开始算的。
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def write_excel(self, row, col, value):
        '''

        :param row: 行数
        :param col: 列数
        :param value:  要填写的值
        :return: 返回值为TRUE或者FALSE，用来判断是否填写成功
        '''
        data = self.open_excel
        wr = data.active
        wr.cell(row, col, value)
        data.save(t_path + '\\excel_data\\excel_data.xlsx')

        # wr.save(t_path + '\\excel_data\\excel_data.xlsx')

    def get_col_value(self,key=None):
        '''

        :param key: 这个可以理解字母为列，同理，数字为行。
        :return: 返回的是整行或者整列的值
        '''
        if key==None:
            key='B'
        col_list=[]
        col_list_data=self.get_sheet_data()[key]
        for col in col_list_data:
            col_list.append(col.value)
        return col_list


    def get_col(self,key,case_id):
        '''

        :param key: 这个可以理解字母为列，同理，数字为行。
        :param case_id: 这个就具体的要查找的值
        :return: 返回行数，第几行
        '''
        num=1
        cols_data=self.get_col_value(key)
        for col_data in cols_data:
            if col_data==case_id:
                return num
            else:
                num+=1

    def get_excel_dict_list(self):
        '''

        :return: 返回的是一个列表嵌套字典的形式
        '''
        row_max=self.get_row()
        key_list=self.get_col_value('2')
        excel_dict_list=[]
        for num in range(row_max-2):
            exce_dict={}
            row_data=self.get_data_row(num+3)
            for i in range(len(row_data)):
                exce_dict[key_list[i]]=row_data[i]
            excel_dict_list.append(exce_dict)
        return excel_dict_list


if __name__ == '__main__':
    re = Read_Excel()
    data=re.get_sheet_data()['A1']
    sheet=re.sheet
    print(data)
    print(sheet)

'''    
    var = re.get_data_row(1)
    #print(var)
    gcv=re.get_col_value('1')
    #num=re.get_col('A','case编号')
    #print(gcv)
    cell_data=re.get_data(2,1)
    #print(cell_data)
    #print(gcv)
    dict_list=re.get_excel_dict_list()
    print(dict_list)
    '''